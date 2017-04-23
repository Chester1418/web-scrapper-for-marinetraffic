# -*- coding: UTF-8 -*-
import os
import time
import urllib
from datetime import datetime
from random import randint
import csv
import psycopg2
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
print ('{0} : start time'.format(datetime.now()))


wb = Workbook()
ws = wb.create_sheet('Sheet')
ws['A1'] = 'mmsidefault'
ws['B1'] = 'exist'
ws['C1'] = 'image'
ws['D1'] = 'type'
ws['E1'] = 'name'
ws['F1'] = 'sort'

filename = 'additional_pasring.xlsx'
list_mmsi = load_workbook(filename)

for name in list_mmsi.get_sheet_names():
    print ('{0} : start time for {1}'.format(datetime.now(),name))
    path = 'C:\Users\User\Desktop\python\parse_marine\{0}'.format(name)
    if not os.path.exists(path):
        os.makedirs(path)
    for i, mmsi in enumerate(list_mmsi.get_sheet_by_name(name)['A'],2):
        try:

            headers = {'Host': 'www.marinetraffic.com',
                       'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0',
                       'Accept': 'text/html, */*; q=0.01',
                       'Accept-Language': 'uk,ru;q=0.8,en-US;q=0.5,en;q=0.3',
                       'Accept-Encoding': 'gzip, deflate, br',
                       'Referer': 'https://www.marinetraffic.com/en/ais/details/ships/mmsi:{0}'.format(mmsi.value),
                       'X-Requested-With': 'XMLHttpRequest',
                       'Cookie': 'SERVERID=www6; vTo=1; CAKEPHP=oc82mpul9s59rvpleefunhhf76; _pendo_meta.915388fa-afe0-454c-6270-7a41b245e92e=736654769; _pendo_visitorId.915388fa-afe0-454c-6270-7a41b245e92e=_PENDO_T_JZ32br29ujf',
                       'DNT': '1',
                       'Connection': 'keep-alive',
                       'Pragma': 'no-cache',
                       'Cache-Control': 'no-cache'}
            link = 'https://www.marinetraffic.com/en/ais/details/ships/mmsi:{0}'.format(mmsi.value)
            page = requests.get(link, headers=headers)
            soup = BeautifulSoup(page.content, 'html.parser')
            soup.prettify()
            r=soup.find_all('div',class_='bg-info')            

            img_path = os.path.join(path, '{0}.jpg'.format(mmsi.value))
	
            ws.cell(column=1, row=i, value=mmsi.value)
            inf = soup.select('div.group-ib.short-line b')
            # saving all data in one row(status, imo,mmsi et)
            ws.cell(column=6, row=i, value=repr(r[0].get_text()))
            #True if page exist for given number
            if inf == []:
                ws.cell(column=2, row=i, value=False)
            else:
                ws.cell(column=2, row=i, value=True)
                if soup.findAll(text='{0}'.format(mmsi.value)) is not None:
                    try:
			#saving image 
                        img = soup.select('a#big-image img')
                        a = img[0].attrs
                        newa = str(a['src'])
                        zzz = 'https:' + newa
                        urllib.urlretrieve(zzz, img_path)
                        ws.cell(column=3, row=i, value=True)
                    except:
                        ws.cell(column=3, row=i, value=False)
                    try:
			# saving type of vessel
                        tp = soup.find_all('a', class_='font-120')
                        name_ship=soup.select('h1.font-200.no-margin')
                        ws.cell(column=4, row=i, value=tp[0].string)
                        ws.cell(column=5, row =i, value=name_ship[0].get_text().replace('\n','|'))
                    except:
                        ws.cell(column=4, row=i,value=' type not found')
                else:
                    ws.cell(column=5, row=i, value='mmsi doesnt match')

            print ('{0} is completed, now sleeep time'.format(i - 1))
            time.sleep(randint(6, 15))
        except:
            print ('{0} :  ERROR time'.format(datetime.now()))
            wb.save('mmsi_{0}_{1}.xlsx'.format(name,datetime.now().strftime('%Y-%m-%d')))
    wb.save('mmsi_{0}.xlsx'.format(name))

    print ('mmsi_{0} was saved'.format(name))
time.sleep(360)
os.system('shutdown -s')
